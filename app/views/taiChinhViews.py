from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from app.models import *
from app.forms import *
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
import urllib
from openpyxl.styles import Alignment
from app.views.utils.blockChain import Blockchain
from django.db.models import ProtectedError


@login_required
def taiChinhHome(request):
    blockChain = Blockchain()

    soBlockInvalid = len(blockChain.is_valid_chain())
    soBlock = len(BlockChain.objects.order_by('index'))-1
    
    soLuongChuaNop = ThanhToanHocPhi.objects.filter(trangThai=0).count()
    soLuongDaNop = ThanhToanHocPhi.objects.filter(trangThai=1).count()
    soTienChuaNop = ThanhToanHocPhi.objects.filter(trangThai=0).aggregate(Sum('soTien'))['soTien__sum']
    soTienDaNop = ThanhToanHocPhi.objects.filter(trangThai=1).aggregate(Sum('soTien'))['soTien__sum']
    return render(request, 'TaiChinh/taiChinhHome.html', {'soLuongChuaNop': soLuongChuaNop,'soLuongDaNop': soLuongDaNop,'soTienChuaNop': soTienChuaNop,'soTienDaNop': soTienDaNop,'soBlockInvalid': soBlockInvalid,'soBlock': soBlock })

@login_required
def XemKhoanThuHocPhi(request,id):
    thanhToanHocPhi = get_object_or_404(ThanhToanHocPhi,id = id)
    return render(request, 'TaiChinh/KhoanThu/xemKhoanThuHocPhi.html', {'thanhToanHocPhi': thanhToanHocPhi})


@login_required
def ListInvalidBlocks(request):
    blockChain = Blockchain()
    invalidBlocks = blockChain.is_valid_chain()
    return render(request, 'TaiChinh/BlockChain/listInvalidBlocks.html', {"invalidBlocks": invalidBlocks})

@login_required
def ListKhoanThuKhac(request):
   alertContent = request.GET.get('alertContent')
   listKhoanThuKhac = KhoanThuKhac.objects.all().order_by('-id')
   return render(request, 'TaiChinh/ThietLapKhoanThuKhac/listKhoanThuKhac.html', {'listKhoanThuKhac': listKhoanThuKhac, 'alertContent': alertContent})

@login_required
def NewKhoanThuKhac(request):
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    cacKhoa= Khoa.objects.all().order_by('-id')
    if request.method == 'POST':
        form = KhoanThuKhacForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listKhoanThuKhac') + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = KhoanThuKhacForm()
    return render(request, 'TaiChinh/ThietLapKhoanThuKhac/newKhoanThuKhac.html', {'form': form, 'cacHeDaoTao': cacHeDaoTao,'cacKhoa': cacKhoa})

@login_required
def SuaKhoanThuKhac(request, id):
    khoanThuKhac = get_object_or_404(KhoanThuKhac, id=id)
    if request.method == 'POST':
        khoanThuKhac.noiDung = request.POST.get('noiDung')
        khoanThuKhac.tenKhoan = request.POST.get('tenKhoan')
        khoanThuKhac.save()
        alert_content = 'Sửa'
        url = reverse('listKhoanThuKhac') + f'?alertContent={alert_content}'
        return redirect(url)
    else:
        khoanThuKhac = KhoanThuKhac.objects.get(id=id)
        form = KhoanThuKhacForm()

    return render(request, 'TaiChinh/ThietLapKhoanThuKhac/suaKhoanThuKhac.html', {'form': form,'khoanThuKhac': khoanThuKhac})

@login_required
def XoaKhoanThuKhac(request, id):
    khoanThuKhac = get_object_or_404(KhoanThuKhac, id=id)
    soTien = format(khoanThuKhac.soTien, ",.0f")
    if request.method == 'POST':
        try:
            khoanThuKhac.delete()
            alert_content = 'Xóa'
            url = reverse('listKhoanThuKhac') + f'?alertContent={alert_content}'
            return redirect(url)
        except ProtectedError:
            return render(request, 'error.html')

    return render(request, 'TaiChinh/ThietLapKhoanThuKhac/xoaKhoanThuKhac.html', {'khoanThuKhac': khoanThuKhac,'soTien': soTien})


@login_required
def ListLop(request):
    ListHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    ListNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    ListKhoa = Khoa.objects.all().order_by('-id')
    ListNamKhoa = [1,2,3,4,5,6,7,8]
    ListLop = Lop.objects.all().order_by('-id')
    namKhoadict = {}

    for nganh_dao_tao in ListNganhDaoTao:
        namKhoa = 0
        id = nganh_dao_tao.id
        for lop in ListLop:
            if lop.nganhDaoTao == nganh_dao_tao and lop.namKhoa > namKhoa:
                namKhoa = lop.namKhoa
        namKhoadict[nganh_dao_tao.id] = namKhoa

    if request.method == 'POST':
        loaiBaoCao = request.POST.get('loaiBaoCao')
        if loaiBaoCao == 'tongHop':
            BaoCaoTongHop()

    return render(request, 'TaiChinh/KhoanThu/chonLop.html', {'ListHeDaoTao': ListHeDaoTao, 'ListNganhDaoTao': ListNganhDaoTao, 'ListKhoa': ListKhoa, 'ListNamKhoa': ListNamKhoa, 'ListLop': ListLop, 'namKhoadict': namKhoadict})

@login_required
def ListSinhVien(request,id):
   lop = get_object_or_404(Lop, id=id)
   ListSinhVien = SinhVien.objects.filter(lop=id).order_by('-id')
   listKhoanThuHocPhi = ThanhToanHocPhi.objects.filter(sinhVien__lop = lop).order_by('-sinhVien__maSV')
   listKhoanThuKhac = ThanhToanKhoanThuKhac.objects.filter(sinhVien__lop = lop).order_by('-sinhVien__maSV')
   alertContent = request.GET.get('alertContent')
   return render(request, 'TaiChinh/KhoanThu/listSinhVien.html', {'listKhoanThuHocPhi': listKhoanThuHocPhi,'ListSinhVien': ListSinhVien,'listKhoanThuKhac': listKhoanThuKhac, 'lop': lop,'alertContent': alertContent,'lopId':id})

#Batch insert 
@login_required
def AddThanhToanKhoanThuKhac(request,id):
    lop = get_object_or_404(Lop, id=id)
    listSinhVien = SinhVien.objects.filter(lop__id = id).order_by('-id')
    listKhoanThuKhac= KhoanThuKhac.objects.filter(Q(khoa=lop.khoa) | Q(khoa=None)).order_by('-id')
    if request.method == 'POST':
        form = BatchThanhToanKhacForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            khoanThuKhacId = request.POST.get('khoanThuKhac')
            khoanThuKhac = get_object_or_404(KhoanThuKhac, id=khoanThuKhacId)
            hanNop  = request.POST.get('hanNop')
            noiDungThu  = request.POST.get('noiDungThu')
            soTien = request.POST.get('soTien').replace(",", "") 
            for sinhVien in listSinhVien:
                khoanThuHocPhi = ThanhToanKhoanThuKhac(sinhVien=sinhVien, khoanThuKhac=khoanThuKhac, hanNop = hanNop , noiDungThu=noiDungThu,
                                                 soTien = soTien, trangThai = 0 )
                khoanThuHocPhi.save()
            alert_content = 'Thêm'
            url = reverse('listSinhVien', args=[id]) + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = BatchThanhToanKhacForm()
    return render(request, 'TaiChinh/KhoanThuKhac/addThanhToanKhac.html', {'form': form, 'listSinhVien': listSinhVien,'listKhoanThuKhac': listKhoanThuKhac,'lopId':id})

@login_required
def NewThanhToanKhoanThuKhac(request,id):
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    cacKhoa= Khoa.objects.all().order_by('-id')
    lop = get_object_or_404(Lop, id=id)
    listSinhVien = SinhVien.objects.filter(lop__id = id).order_by('-id')
    listKhoanThuKhac= KhoanThuKhac.objects.filter(Q(khoa=lop.khoa) | Q(khoa=None)).order_by('-id')
    if request.method == 'POST':
        form = ThanhToanKhacForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listSinhVien', args=[id]) + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = ThanhToanKhacForm()
    return render(request, 'TaiChinh/KhoanThuKhac/newThanhToanKhac.html', {'form': form, 'listSinhVien': listSinhVien,'listKhoanThuKhac': listKhoanThuKhac,'lopId':id,'cacKhoa':cacKhoa,'cacHeDaoTao':cacHeDaoTao})


# @login_required
# def SuaThanhToanKhoanThuKhac(request, id):
    
#     thanhToanKhac = ThanhToanKhoanThuKhac.objects.get(id=id)
#     listKhoanThuKhac= KhoanThuKhac.objects.filter(lop = thanhToanKhac.sinhVien.lop).order_by('-id')
#     if request.method == 'POST':
#         thanhToanKhac.sinhVien = get_object_or_404(SinhVien, id=request.POST.get('sinhVien'))
#         thanhToanKhac.khoanThuKhac = get_object_or_404(KhoanThuKhac, id=request.POST.get('khoanThuKhac'))
#         thanhToanKhac.hanNop = request.POST.get('hanNop')
#         thanhToanKhac.noiDungThu = request.POST.get('noiDungThu')
#         thanhToanKhac.ghiChu = request.POST.get('ghiChu')
#         thanhToanKhac.soTien = request.POST.get('soTien').replace(",", "")

#         thanhToanKhac.save()
#         alert_content = 'Sửa'
#         url = reverse('listSinhVien', args=[thanhToanKhac.sinhVien.lop.id]) + f'?alertContent={alert_content}'
#         return redirect(url)
#     else:
        
#         form = ThanhToanKhacForm()

#     return render(request, 'TaiChinh/KhoanThuKhac/suaThanhToanKhac.html', {'form': form,'thanhToanKhac': thanhToanKhac,'listKhoanThuKhac': listKhoanThuKhac})

@login_required
def XoaThanhToanKhoanThuKhac(request, id):
    thanhToanKhac = get_object_or_404(ThanhToanKhoanThuKhac, id=id)
    lopId = thanhToanKhac.sinhVien.lop.id

    if request.method == 'POST':
        thanhToanKhac.delete()
        alert_content = 'Xóa'
        url = reverse('listSinhVien', args=[thanhToanKhac.sinhVien.lop.id]) + f'?alertContent={alert_content}'
        return redirect(url)

    return render(request, 'TaiChinh/KhoanThuKhac/xoaThanhToanKhac.html', {'thanhToanKhac': thanhToanKhac,'lopId':lopId})

@login_required
def XemThanhToanKhoanThuKhac(request, id):
    thanhToanKhac = get_object_or_404(ThanhToanKhoanThuKhac, id=id)
    return render(request, 'TaiChinh/KhoanThuKhac/xemThanhToanKhac.html', {'thanhToanKhac': thanhToanKhac})


@login_required
def XemCacKhoanThuCuaSV(request,id):
   sinhVien = get_object_or_404(SinhVien, id=id)
   listKhoanHocPhi = ThanhToanHocPhi.objects.filter(sinhVien=sinhVien).order_by('-id')
   listKhoanThuKhac = ThanhToanKhoanThuKhac.objects.filter(sinhVien=sinhVien).order_by('-id')
   return render(request, 'TaiChinh/KhoanThu/xemCacKhoanThuCuaSV.html', {'listKhoanHocPhi': listKhoanHocPhi,'listKhoanThuKhac': listKhoanThuKhac,'sinhVien': sinhVien})


@login_required
def ChonBaoCao(request):
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacKhoa = Khoa.objects.all().order_by('-id')
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacNamHoc = NamHoc.objects.all().order_by('-id')
    cacLop = Lop.objects.all().order_by('-id')
    cacLoaiBaoCao = ["Báo cáo học phí tổng hợp","Báo cáo học phí tổng hợp theo lớp","Báo cáo công nợ học phí sinh viên","Báo cáo công nợ học phí sinh viên theo lớp"]

    if request.method == 'POST':
        loaiBaoCao = request.POST.get('loaiBaoCao')
        namBatDau = request.POST.get('namBatDau')
        namKetThuc = request.POST.get('namKetThuc')
        if loaiBaoCao == 'Báo cáo học phí tổng hợp':
            return BaoCaoTongHop(namBatDau,namKetThuc)
        elif loaiBaoCao == 'Báo cáo học phí tổng hợp theo lớp':
            lop = request.POST.get('lop')
            return BaoCaoTongHopTheoLop(lop,namBatDau,namKetThuc)
        elif loaiBaoCao == 'Báo cáo công nợ học phí sinh viên':
            return BaoCaoCongNoSinhVien(namBatDau,namKetThuc)
        elif loaiBaoCao == 'Báo cáo công nợ học phí sinh viên theo lớp':
            lop = request.POST.get('lop')
            return BaoCaoCongNoTheoLop(lop,namBatDau,namKetThuc)

    return render(request, 'TaiChinh/BaoCaoThu/chonBaoCao.html', {'cacHeDaoTao': cacHeDaoTao, 'cacNganhDaoTao': cacNganhDaoTao, 'cacKhoa': cacKhoa, 'cacNamHoc': cacNamHoc, 'cacKhoaHoc': cacKhoaHoc, 'cacLop': cacLop, 'cacLoaiBaoCao': cacLoaiBaoCao})

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from django.http import HttpResponse
import urllib.parse

def BaoCaoTongHop(namBatDau,namKetThuc):
    ListKhoa = Khoa.objects.all().order_by('-id')
    workbook = Workbook()
    worksheet = workbook.active
    namHocKetXuat =  namBatDau + "-" + namKetThuc
    start_range = f"{namBatDau}-{int(namBatDau) + 1}"
    end_range = f"{int(namKetThuc) - 1}-{namKetThuc}"
    for khoa in ListKhoa:
        queryset = ThanhToanHocPhi.objects.filter(sinhVien__lop__khoa=khoa).filter(hocPhanDuocXetDuyet__hocPhanDaDangKy__namHoc__namHoc__range=(start_range, end_range)).order_by('sinhVien__maSV')
        worksheet = workbook.create_sheet(title=khoa.tenKhoa)
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if obj.hocPhanDuocXetDuyet.hocPhanDaDangKy.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                obj.hocPhanDuocXetDuyet.hocPhanDaDangKy.namHoc.namHoc,
                hocKy,
                format(obj.soTien, ","),
                trangThai

            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = length + 2

        worksheet.insert_rows(1)
        worksheet['A1'] = 'Báo cáo tổng hợp học phí khoa ' + khoa.tenKhoa + ' - Năm học ' + namHocKetXuat
        worksheet.merge_cells('A1:H1')
        merged_cell = worksheet['A1']
        merged_cell.alignment = Alignment(horizontal='center')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = urllib.parse.quote('Báo cáo tổng hợp.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response


def BaoCaoTongHopTheoLop(lopId,namBatDau,namKetThuc):
    lopDuocChon = get_object_or_404(Lop, id=lopId)
    namHocKetXuat =  namBatDau + "-" + namKetThuc
    start_range = f"{namBatDau}-{int(namBatDau) + 1}"
    end_range = f"{int(namKetThuc) - 1}-{namKetThuc}"
    listNamHoc = NamHoc.objects.filter(namHoc__range=(start_range, end_range)).order_by('-id')
    name = 'Báo cáo tổng hợp theo lớp '+ lopDuocChon.lop + ' - Năm học '+ namHocKetXuat +'.xlsx'
    filename = urllib.parse.quote(name)
    workbook = Workbook()
    worksheet = workbook.active
    for namHoc in listNamHoc:
        queryset = ThanhToanHocPhi.objects.filter(Q(hocPhanDuocXetDuyet__hocPhanDaDangKy__namHoc=namHoc) & Q(hocPhanDuocXetDuyet__hocPhanDaDangKy__hocKy=1) & Q(sinhVien__lop=lopDuocChon)).order_by('sinhVien__maSV')
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 1")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            thongTin = obj.hocPhanDuocXetDuyet.hocPhanDaDangKy
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if thongTin.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                thongTin.namHoc.namHoc,
                hocKy,
                format(obj.soTien, ","),
                trangThai
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

        worksheet.insert_rows(1)
        worksheet['A1'] = 'Báo cáo tổng hợp học phí lớp ' + lopDuocChon.lop + " - Năm học "+ namHoc.namHoc + " - Học kỳ 1"
        worksheet.merge_cells('A1:H1')
        merged_cell = worksheet['A1']
        merged_cell.alignment = Alignment(horizontal='center')

        queryset = ThanhToanHocPhi.objects.filter(Q(hocPhanDuocXetDuyet__hocPhanDaDangKy__namHoc=namHoc) & Q(hocPhanDuocXetDuyet__hocPhanDaDangKy__hocKy=0) & Q(sinhVien__lop=lopDuocChon)).order_by('sinhVien__maSV')
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 2")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            thongTin = obj.hocPhanDuocXetDuyet.hocPhanDaDangKy
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if thongTin.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                thongTin.namHoc.namHoc,
                hocKy,
                format(obj.soTien, ","),
                trangThai
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

        worksheet.insert_rows(1)
        worksheet['A1'] = 'Báo cáo tổng hợp học phí lớp ' + lopDuocChon.lop + " - Năm học "+ namHoc.namHoc + " - Học kỳ 2"
        worksheet.merge_cells('A1:H1')
        merged_cell = worksheet['A1']
        merged_cell.alignment = Alignment(horizontal='center')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response

def BaoCaoCongNoSinhVien(namBatDau,namKetThuc):
    ListKhoa = Khoa.objects.all().order_by('-id')
    namHocKetXuat =  namBatDau + "-" + namKetThuc
    start_range = f"{namBatDau}-{int(namBatDau) + 1}"
    end_range = f"{int(namKetThuc) - 1}-{namKetThuc}"
    workbook = Workbook()
    worksheet = workbook.active
    for khoa in ListKhoa:
        queryset = ThanhToanHocPhi.objects.filter(sinhVien__lop__khoa=khoa).filter(hocPhanDuocXetDuyet__hocPhanDaDangKy__namHoc__namHoc__range=(start_range, end_range)).order_by('sinhVien__maSV').filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('soTien'))
        worksheet = workbook.create_sheet(title=khoa.tenKhoa)
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']

        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2
        
        worksheet.insert_rows(1)
        worksheet['A1'] = 'Báo cáo tổng hợp công nợ học phí sinh viên khoa ' + khoa.tenKhoa + ' - Năm học ' + namHocKetXuat
        worksheet.merge_cells('A1:E1')
        merged_cell = worksheet['A1']
        merged_cell.alignment = Alignment(horizontal='center')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = urllib.parse.quote('Báo cáo công nợ sinh viên - Năm học '+ namHocKetXuat +'.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response

def BaoCaoCongNoTheoLop(lopId,namBatDau,namKetThuc):
    namHocKetXuat =  namBatDau + "-" + namKetThuc
    start_range = f"{namBatDau}-{int(namBatDau) + 1}"
    end_range = f"{int(namKetThuc) - 1}-{namKetThuc}"
    lopDuocChon = get_object_or_404(Lop, id=lopId)
    listNamHoc = NamHoc.objects.filter(namHoc__range=(start_range, end_range)).order_by('-id')
    name = 'Báo cáo công nợ sinh viên theo lớp '+ lopDuocChon.lop + ' - Năm học '+ namHocKetXuat +'.xlsx'
    filename = urllib.parse.quote(name)
    workbook = Workbook()
    worksheet = workbook.active
    for namHoc in listNamHoc:
        queryset = ThanhToanHocPhi.objects.filter(Q(hocPhanDuocXetDuyet__hocPhanDaDangKy__namHoc=namHoc) & Q(hocPhanDuocXetDuyet__hocPhanDaDangKy__hocKy=1) & Q(sinhVien__lop=lopDuocChon)).order_by('sinhVien__maSV').filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('soTien'))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 1")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

        worksheet.insert_rows(1)
        worksheet['A1'] = 'Báo cáo công nợ học phí lớp '+ lopDuocChon.lop + " - Năm học "+ namHoc.namHoc + " - Học kỳ 1"
        worksheet.merge_cells('A1:E1')
        merged_cell = worksheet['A1']
        merged_cell.alignment = Alignment(horizontal='center')

        queryset = ThanhToanHocPhi.objects.filter(Q(hocPhanDuocXetDuyet__hocPhanDaDangKy__namHoc=namHoc) & Q(hocPhanDuocXetDuyet__hocPhanDaDangKy__hocKy=0) & Q(sinhVien__lop=lopDuocChon)).order_by('sinhVien__maSV').filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('soTien'))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 2")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

        worksheet.insert_rows(1)
        worksheet['A1'] = 'Báo cáo công nợ học phí lớp '+ lopDuocChon.lop + " - Năm học "+ namHoc.namHoc + " - Học kỳ 2"
        worksheet.merge_cells('A1:E1')
        merged_cell = worksheet['A1']
        merged_cell.alignment = Alignment(horizontal='center')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response

@login_required
def ListKeHoachThu(request):
   alertContent = request.GET.get('alertContent')
   listKeHoachThu = KeHoachThu.objects.all().order_by('-id')
   return render(request, 'TaiChinh/KeHoachThu/listKeHoachThu.html', {'listKeHoachThu': listKeHoachThu, 'alertContent': alertContent})

@login_required
def NewKeHoachThu(request):
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    if request.method == 'POST':
        form = KeHoachThuForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listKeHoachThu') + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = KeHoachThuForm()
    return render(request, 'TaiChinh/KeHoachThu/newKeHoachThu.html', {'form': form, 'cacHeDaoTao': cacHeDaoTao,'cacKhoaHoc': cacKhoaHoc,'cacNganhDaoTao': cacNganhDaoTao})

@login_required
def SuaKeHoachThu(request, id):

    keHoachThu = get_object_or_404(KeHoachThu, id=id)
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    if request.method == 'POST':
        keHoachThu.khoaHoc =  get_object_or_404(KhoaHoc, id=request.POST.get('khoaHoc'))
        keHoachThu.khoanThu = request.POST.get('khoanThu')
        keHoachThu.nganhDaoTao = get_object_or_404(NganhDaoTao, id=request.POST.get('nganhDaoTao'))
        keHoachThu.soTien = request.POST.get('soTien').replace(",", "")
        keHoachThu.dotThu = request.POST.get('dotThu')
        keHoachThu.ngayBatDau = request.POST.get('ngayBatDau')
        keHoachThu.ngayKetThuc = request.POST.get('ngayKetThuc')
        keHoachThu.noiDung = request.POST.get('noiDung')
        keHoachThu.save()
        alert_content = 'Sửa'
        url = reverse('listKeHoachThu') + f'?alertContent={alert_content}'
        return redirect(url)
    else:
        keHoachThu = KeHoachThu.objects.get(id=id)
        form = KeHoachThuForm()


    return render(request, 'TaiChinh/KeHoachThu/suaKeHoachThu.html', {'form': form,'keHoachThu': keHoachThu,'cacHeDaoTao': cacHeDaoTao,'cacKhoaHoc': cacKhoaHoc,'cacNganhDaoTao': cacNganhDaoTao})

@login_required
def XoaKeHoachThu(request, id):
    keHoachThu = get_object_or_404(KeHoachThu, id=id)
    soTien = format(keHoachThu.soTien, ",.0f")
    if request.method == 'POST':
        keHoachThu.delete()
        alert_content = 'Xóa'
        url = reverse('listKeHoachThu') + f'?alertContent={alert_content}'
        return redirect(url)

    return render(request, 'TaiChinh/KeHoachThu/xoaKeHoachThu.html', {'keHoachThu': keHoachThu, 'soTien': soTien})

@login_required
def XemKeHoachThu(request, id):
    keHoachThu = get_object_or_404(KeHoachThu, id=id)
    soTien = format(keHoachThu.soTien, ",.0f")
    return render(request, 'TaiChinh/KeHoachThu/xemKeHoachThu.html', {'keHoachThu': keHoachThu, 'soTien': soTien})


@login_required
def ListLichSuThanhToan(request,id):
   sinhVien = get_object_or_404(SinhVien, id=id)
   listLichSuHocPhi = BlockChain.objects.filter(maSinhVien = sinhVien.maSV, maThanhToan__icontains="HP").order_by('-id')
   listLichSuKhac = BlockChain.objects.filter(maSinhVien = sinhVien.maSV, maThanhToan__icontains="K").order_by('-id')
   return render(request, 'TaiChinh/KhoanThu/listLichSuThanhToan.html', {'listLichSuHocPhi': listLichSuHocPhi,'listLichSuKhac': listLichSuKhac,'sinhVien': sinhVien})



